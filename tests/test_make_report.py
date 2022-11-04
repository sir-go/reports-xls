import hashlib
import json

import openpyxl
import pytest

from make_report import make_report, make_cities_report
from openpyxl import Workbook


@pytest.fixture
def db_data():
    with open('tests/dumps/report.json', 'r') as dfd:
        return json.loads(dfd.read())


def workbook_md5(wb: Workbook):
    mds = hashlib.md5()
    for ws in wb.worksheets:
        for row in ws.rows:
            for cell in row:
                mds.update(bytes(str(cell.value), 'utf-8'))
    return mds.hexdigest()


def test_make_report(db_data, tmp_path):
    wb_test = openpyxl.load_workbook('tests/dumps/report_10.01.2048.xlsx')

    wb = Workbook()
    ws_first = wb.active
    ws_first.title = u'ФЛ по нас.пунктам'
    today = '10.01.2048'
    data = db_data
    make_cities_report(today, data=data, jur=False, ws=ws_first)
    make_cities_report(today, data=data, jur=True,
                       ws=wb.create_sheet(u'ЮЛ по нас.пунктам'))
    make_report(today, data=data, jur=False, ws=wb.create_sheet(u'ФЛ все'))
    make_report(today, data=data, jur=True, ws=wb.create_sheet(u'ЮЛ все'))

    assert workbook_md5(wb_test) == workbook_md5(wb)
