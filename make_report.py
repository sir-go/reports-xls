# coding:utf-8
import os
from itertools import groupby
from openpyxl import Workbook
from openpyxl.worksheet import worksheet
from openpyxl.styles import Font, Alignment, PatternFill, fills
from openpyxl.utils import get_column_letter
import pymysql
import pymysql.cursors
from os import path, environ
from time import strftime

conf = dict(
    host=environ.get('REP_HOST'),
    user=environ.get('REP_USERNAME'),
    passwd=environ.get('REP_PASSWORD'),
    db=environ.get('REP_DB')
)


def get_from_db():
    qry = '''
    select
      s.uid,
      s.jur,
      s.house_id,
      s.city,
      s.addr,
      s.tname,
      s.speed,
      s.price
    from
      stat_tariffs_addresses as s
    where
      s.date = current_date
    group by uid
    order by addr   
    '''
    connection = pymysql.connect(**conf,
                                 use_unicode=True,
                                 charset="utf8",
                                 cursorclass=pymysql.cursors.DictCursor,
                                 autocommit=True)
    with connection:
        with connection.cursor() as cursor:
            cursor.execute(qry)
            q_res = cursor.fetchall()
    return q_res


def make_report(
        today: str, data: list[dict], jur: bool = False, ws: worksheet = None):
    # filter by juridical condition

    data_ = [record for record in data if record['jur'] == jur]

    # styles
    vertical_middle = Alignment(vertical='center')
    horizontal_center_90_deg = Alignment(horizontal='center', text_rotation=90)
    font_arial_10 = Font(name='Arial', sz=10)
    font_arial_12_bold = Font(name='Arial', sz=12, b=True)

    # header

    title = u'Отёт по тарифам {} {}'.format(
        u'ЮЛ' if jur else u'ФЛ',
        today
    )

    ws['A1'] = title

    ws['A2'] = u'наименование  тарифа'
    ws['A2'].alignment = vertical_middle

    ws['A3'] = u'скорость Мбит/с'
    ws['A4'] = u'цена р/мес'
    ws['A5'] = u'кол-во абонентов'
    ws['A6'] = u'доходность р/мес'

    # tariffs row

    tariffs = []

    for rec in data_:
        speed = int(rec['speed'])
        cost = int(rec['price'])

        tariffs.append({
            'speed': speed,
            'cost': cost,
            'name': rec['tname']
        })

    added_tariffs = {}
    col = 2
    for t in sorted(tariffs, key=lambda tr: (tr['speed'], tr['cost'])):
        if t['name'] not in added_tariffs:
            added_tariffs[t['name']] = col
            ws.cell(column=col, row=2, value=t['name'])
            ws.cell(column=col, row=3, value=t['speed'])
            ws.cell(column=col, row=4, value=t['cost'])
            col += 1
    ws.cell(column=col, row=2, value=u"СУММА")

    # addresses column

    matrix = {}
    row = 7
    added_addresses = {}
    for r in sorted(data_, key=lambda x: x['addr']):

        hash_ = hash(r['addr']) + hash(r['tname'])
        if hash_ in matrix:
            matrix[hash_] += 1
        else:
            matrix[hash_] = 1

        if r['addr'] not in added_addresses:
            added_addresses[r['addr']] = row
            ws.cell(column=1, row=row, value=r['addr'])
            row += 1

    # fill the table

    for record in data_:
        count = matrix[hash(record['addr']) + hash(record['tname'])]
        col = added_tariffs[record['tname']]
        row = added_addresses[record['addr']]
        ws.cell(column=col, row=row, value=count)

    # sum row

    for cur_col in range(2, len(added_tariffs) + 2):
        cur_row = 5
        cell_first = ws.cell(column=cur_col, row=cur_row + 2)
        cell_last = ws.cell(column=cur_col,
                            row=cur_row + len(added_addresses) + 1)

        formula_sum = '=SUM({}:{})'.format(
            cell_first.coordinate, cell_last.coordinate)
        ws.cell(column=cur_col, row=cur_row, value=formula_sum)

        cell_amount = ws.cell(column=cur_col, row=cur_row)
        cell_price = ws.cell(column=cur_col, row=cur_row - 1)
        formula_profit = '={}*{}'.format(
            cell_amount.coordinate, cell_price.coordinate)
        ws.cell(column=cur_col, row=cur_row + 1, value=formula_profit)

    # # sum column
    #

    cur_row = 7
    for row in range(cur_row - 2, len(added_addresses) + cur_row):
        cf = ws.cell(column=2, row=row)
        ce = ws.cell(column=len(added_tariffs) + 1, row=row)

        # print(cf.coordinate, ce.coordinate)

        formula_sum = '=SUM({}:{})'.format(cf.coordinate, ce.coordinate)
        ws.cell(column=len(added_tariffs) + 2, row=row, value=formula_sum)

    # freeze

    ws.freeze_panes = ws['B7']

    # rotate tariffs cells

    for col in range(2, ws.max_column + 1):
        c = ws.cell(column=col, row=2)
        c.alignment = horizontal_center_90_deg
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 7

    # 1st column width, tariffs row height

    ws.column_dimensions['A'].width = 35
    ws.row_dimensions[2].height = 170

    # apply font styles

    for row in ws.iter_rows():
        for cell in row:
            cell.font = font_arial_10
    ws['A1'].font = font_arial_12_bold


def make_cities_report(
        today: str, data: list[dict], jur: bool = False, ws: worksheet = None):
    # filter by juridical condition

    data_ = [record for record in data if record['jur'] == jur]

    # styles
    vertical_middle = Alignment(vertical='center')
    horizontal_center_90_deg = Alignment(horizontal='center', text_rotation=90)
    style_title = Alignment(wrap_text=True)
    font_arial_10 = Font(name='Arial', sz=10)
    font_arial_12_bold = Font(name='Arial', sz=12, b=True)
    font_arial_10_bold = Font(name='Arial', sz=10, b=True)
    # noinspection SpellCheckingInspection
    grayed_style = PatternFill(
        fill_type=fills.FILL_SOLID, start_color='eeeeee')

    # header

    title = u'Отчёт по тарифам {} в разрезе населенных пунктов на {}'.format(
        u'ЮЛ' if jur else u'ФЛ',
        today
    )

    ws['A1'] = title
    ws['A1'].alignment = style_title
    ws.row_dimensions[1].height = 65

    ws['A2'] = u'наименование  тарифа'
    ws['A2'].alignment = vertical_middle

    ws['A3'] = u'скорость Мбит/с'
    ws['A4'] = u'цена р/мес'
    price_row = 4

    ws['A5'] = u'кол-во абонентов'
    amount_row = 5
    amount_cells = {}

    ws['A6'] = u'доходность р/мес'
    profit_row = 6
    profit_cells = {}

    grid_begin_row = profit_row + 1

    # tariffs row

    tariffs = []

    for rec in data_:
        speed = int(rec['speed'])
        cost = int(rec['price'])

        tariffs.append({
            'speed': speed,
            'cost': cost,
            'name': rec['tname']
        })

    added_tariffs = {}
    col = 3
    for t in sorted(tariffs, key=lambda tr: (tr['speed'], tr['cost'])):
        if t['name'] not in added_tariffs:
            added_tariffs[t['name']] = col
            ws.cell(column=col, row=2, value=t['name'])
            ws.cell(column=col, row=3, value=t['speed'])
            ws.cell(column=col, row=4, value=t['cost'])
            col += 1

    ws.freeze_panes = ws['C7']

    # rotate tariffs cells

    for col in range(3, ws.max_column + 1):
        c = ws.cell(column=col, row=2)
        c.alignment = horizontal_center_90_deg
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 7

    # 1st column width, tariffs row height

    ws.column_dimensions['A'].width = 25
    ws.row_dimensions[2].height = 170

    cur_row = 7
    cur_col = 1
    grouped_by_cities = groupby(data_, lambda x: x['city'])
    for city, records in grouped_by_cities:
        cur_cell = ws.cell(column=cur_col, row=cur_row, value=city)
        ws.merge_cells(
            start_row=cur_row, start_column=cur_col,
            end_row=cur_row + 1, end_column=cur_col
        )
        cur_cell.alignment = vertical_middle
        ws.cell(column=cur_col + 1, row=cur_row, value=u"кол-во")
        ws.cell(column=cur_col + 1, row=cur_row + 1, value=u"доходность")

        for tariff_name, tariff_column in added_tariffs.items():
            records_with_this_tariff = [
                r for r in data_
                if r['city'] == city and r['tname'] == tariff_name]
            cur_count_cell = ws.cell(
                column=tariff_column, row=cur_row,
                value=len(records_with_this_tariff))

            price_cell = ws.cell(column=tariff_column, row=price_row)
            cur_profit_cell = ws.cell(
                column=tariff_column, row=cur_row + 1, value='={}*{}'.format(
                    price_cell.coordinate,
                    cur_count_cell.coordinate
                ))

            if tariff_column not in amount_cells:
                amount_cells[tariff_column] = []

            amount_cells[tariff_column].append(cur_count_cell.coordinate)

            if tariff_column not in profit_cells:
                profit_cells[tariff_column] = []

            profit_cells[tariff_column].append(cur_profit_cell.coordinate)

        cf = ws.cell(column=cur_col + 2, row=cur_row)
        ce = ws.cell(column=len(added_tariffs) + 2, row=cur_row)
        ws.cell(column=len(added_tariffs) + 3, row=cur_row,
                value='=SUM({}:{})'.format(cf.coordinate, ce.coordinate))

        cf = ws.cell(column=cur_col + 2, row=cur_row + 1)
        ce = ws.cell(column=len(added_tariffs) + 2, row=cur_row + 1)
        ws.cell(column=len(added_tariffs) + 3, row=cur_row + 1,
                value='=SUM({}:{})'.format(cf.coordinate, ce.coordinate))

        cur_row += 2

    # sum row
    sum_amount_cf = ws.cell(column=cur_col + 2, row=amount_row)
    sum_amount_ce = ws.cell(column=len(added_tariffs) + 2, row=amount_row)
    sum_amount_cell = ws.cell(column=len(added_tariffs) + 3, row=amount_row)
    sum_amount_cell.value = '=SUM({}:{})'.format(sum_amount_cf.coordinate,
                                                 sum_amount_ce.coordinate)

    sum_profit_cf = ws.cell(column=cur_col + 2, row=profit_row)
    sum_profit_ce = ws.cell(column=len(added_tariffs) + 2, row=profit_row)
    sum_profit_cell = ws.cell(column=len(added_tariffs) + 3, row=profit_row)
    sum_profit_cell.value = '=SUM({}:{})'.format(sum_profit_cf.coordinate,
                                                 sum_profit_ce.coordinate)

    # profit & amount sums row

    for clm, cells in amount_cells.items():
        ws.cell(column=clm, row=amount_row,
                value='={}'.format('+'.join(cells)))

    for clm, cells in profit_cells.items():
        ws.cell(column=clm, row=profit_row,
                value='={}'.format('+'.join(cells)))

    # apply font styles
    sum_cell = ws.cell(column=len(added_tariffs) + 3, row=amount_row)
    strip_counter = 1

    for idx, row in enumerate(ws.iter_rows()):
        if idx >= grid_begin_row:
            strip_counter += 1
        for c_idx, cell in enumerate(row):
            if c_idx > 1 and not (strip_counter % 2):
                cell.fill = grayed_style
            cell.font = font_arial_10
            if cell.column == sum_cell.column:
                cell.font = font_arial_10_bold

    ws['A1'].font = font_arial_12_bold


def make_all_reports():
    wb = Workbook()
    ws_first = wb.active
    ws_first.title = u'ФЛ по нас.пунктам'
    today = strftime('%d.%m.%Y')
    data = get_from_db()
    make_cities_report(today, data=data, jur=False,
                       ws=ws_first)
    make_cities_report(today, data=data, jur=True,
                       ws=wb.create_sheet(u'ЮЛ по нас.пунктам'))
    make_report(today, data=data, jur=False, ws=wb.create_sheet(u'ФЛ все'))
    make_report(today, data=data, jur=True, ws=wb.create_sheet(u'ЮЛ все'))

    here = path.dirname(path.abspath(__file__))
    filename = path.join(os.environ.get('REP_OUT', here),
                         'report_{}.xlsx'.format(strftime('%Y-%m-%d')))

    wb.save(filename)


if __name__ == '__main__':
    make_all_reports()
